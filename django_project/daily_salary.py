import pandas as pd
from datetime import datetime, timedelta
import holidays
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Constants
HOURLY_RATE = 24.0385
HOURS_PER_DAY = 8
POSITION = 'Data Analyst'
us_holidays = holidays.US()

def is_business_day(date):
    return date.weekday() < 5 and date not in us_holidays

def generate_data(start_date, end_date):
    current_date = start_date
    data = []
    while current_date <= end_date:
        if is_business_day(current_date):
            daily_salary = HOURLY_RATE * HOURS_PER_DAY
            data.append({
                'Date': current_date.strftime('%m-%d-%Y'),
                'Position': POSITION,
                'Hourly Salary': HOURLY_RATE,
                'Hours Worked': HOURS_PER_DAY,
                'Daily Salary': daily_salary
            })
        current_date += timedelta(days=1)
    return pd.DataFrame(data)

def save_to_excel(df, output_path):
    total_salary = df['Daily Salary'].sum()
    # Add a total row at the end
    total_row = {'Date': 'Total', 'Position': '', 'Hourly Salary': '', 'Hours Worked': '', 'Daily Salary': total_salary}
    df = df._append(total_row, ignore_index=True)
    
    # Create a workbook and select active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Adding a title
    ws.append(['Salary Report'])
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Adding column headers
    ws.append(['Date', 'Position', 'Hourly Salary', 'Hours Worked', 'Daily Salary'])
    for cell in ws[2]:
        cell.font = Font(bold=True)
    
    # Append dataframe rows
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    
    # Apply font size to table rows
    for row in ws.iter_rows(min_row=3, max_row=len(df) + 3, min_col=1, max_col=5):
        for cell in row:
            cell.font = Font(size=12)
    
    # Save the workbook
    wb.save(output_path)

if __name__ == "__main__":
    # Fetch the output directory from the database
    config = PathConfig.objects.get(name="Default")
    output_directory = config.output_directory  # Use the directory from the database
    
    # Ensure this directory exists
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Define the date range
    start_date = datetime(2023, 1, 1)
    end_date = datetime(2023, 1, 31)
    
    # Generate data and save to Excel
    df = generate_data(start_date, end_date)
    today_str = datetime.now().strftime('%m-%d-%Y')
    filename = f"{today_str} Salary Compensation.xlsx"
    output_path = os.path.join(output_directory, filename)
    save_to_excel(df, output_path)
    print(f"Excel file saved as: {output_path}")
